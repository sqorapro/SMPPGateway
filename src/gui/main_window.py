import sys
import os

if getattr(sys, 'frozen', False):
    _src_path = os.path.join(sys._MEIPASS, 'src')
else:
    _src_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _src_path)

import threading
import time
import subprocess
import tempfile
import logging

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QFileDialog, QLineEdit, QTextEdit, QSpinBox, QGroupBox, QFormLayout, QGridLayout, QDateTimeEdit,
    QCheckBox, QMessageBox, QProgressBar, QComboBox, QSplitter, QStatusBar,
    QFrame, QScrollArea, QSizePolicy, QListWidget, QListWidgetItem
)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QObject, QDateTime
from PyQt5.QtGui import QFont, QColor, QIcon, QPalette

from modem_manager import ModemManager
from database import Database
from smpp_server import SMPPServer, get_local_ip
from call_handler import CallHandler
from sms_handler import SmsHandler


# Параметры VPS для публикации SMPP в интернет (reverse-туннель)
# Заполните своими значениями или вынесите в переменные окружения / config-файл
TUNNEL_VPS_HOST = os.environ.get("SMPP_TUNNEL_VPS_HOST", "your-vps-ip")
TUNNEL_VPS_USER = os.environ.get("SMPP_TUNNEL_VPS_USER", "your-vps-user")
TUNNEL_DOMAIN = os.environ.get("SMPP_TUNNEL_DOMAIN", "your-domain.example")


STYLESHEET = """
QMainWindow {
    background-color: #1e1e2e;
}
QWidget {
    background-color: #1e1e2e;
    color: #cdd6f4;
    font-family: 'Segoe UI', 'SF Pro Display', sans-serif;
    font-size: 13px;
}
QTabWidget::pane {
    border: 1px solid #313244;
    border-radius: 8px;
    background-color: #181825;
    margin-top: -1px;
}
QTabBar::tab {
    background-color: #181825;
    color: #6c7086;
    padding: 10px 28px;
    min-width: 90px;
    border: 1px solid #313244;
    border-bottom: none;
    border-top-left-radius: 8px;
    border-top-right-radius: 8px;
    margin-right: 2px;
    font-weight: 500;
}
QTabBar::tab:selected {
    background-color: #1e1e2e;
    color: #89b4fa;
    border-color: #89b4fa;
    border-bottom: 2px solid #89b4fa;
}
QTabBar::tab:hover:!selected {
    background-color: #313244;
    color: #cdd6f4;
}
QGroupBox {
    border: 1px solid #313244;
    border-radius: 8px;
    margin-top: 16px;
    padding: 16px 12px 12px 12px;
    background-color: #181825;
    font-weight: 600;
    color: #89b4fa;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 14px;
    padding: 0 6px;
}
QPushButton {
    background-color: #313244;
    color: #cdd6f4;
    border: 1px solid #45475a;
    border-radius: 6px;
    padding: 8px 18px;
    font-weight: 500;
}
QPushButton:hover {
    background-color: #45475a;
    border-color: #585b70;
}
QPushButton:pressed {
    background-color: #181825;
}
QPushButton:disabled {
    background-color: #181825;
    color: #45475a;
    border-color: #313244;
}
QPushButton#startBtn {
    background-color: #a6e3a1;
    color: #1e1e2e;
    border: none;
    font-weight: 700;
    font-size: 14px;
    padding: 10px 24px;
}
QPushButton#startBtn:hover {
    background-color: #94d689;
}
QPushButton#startBtn:disabled {
    background-color: #313244;
    color: #45475a;
}
QPushButton#stopBtn {
    background-color: #f38ba8;
    color: #1e1e2e;
    border: none;
    font-weight: 700;
    font-size: 14px;
    padding: 10px 24px;
}
QPushButton#stopBtn:hover {
    background-color: #eb7693;
}
QPushButton#stopBtn:disabled {
    background-color: #313244;
    color: #45475a;
}
QPushButton#smppStartBtn {
    background-color: #a6e3a1;
    color: #1e1e2e;
    border: none;
    font-weight: 700;
    padding: 10px 24px;
}
QPushButton#smppStartBtn:hover {
    background-color: #94d689;
}
QPushButton#smppStartBtn:disabled {
    background-color: #313244;
    color: #45475a;
}
QPushButton#smppStopBtn {
    background-color: #f38ba8;
    color: #1e1e2e;
    border: none;
    font-weight: 700;
    padding: 10px 24px;
}
QPushButton#smppStopBtn:hover {
    background-color: #eb7693;
}
QPushButton#smppStopBtn:disabled {
    background-color: #313244;
    color: #45475a;
}
QLineEdit, QSpinBox, QComboBox {
    background-color: #313244;
    color: #cdd6f4;
    border: 1px solid #45475a;
    border-radius: 6px;
    padding: 6px 10px;
    selection-background-color: #89b4fa;
    selection-color: #1e1e2e;
}
QLineEdit:focus, QSpinBox:focus, QComboBox:focus {
    border-color: #89b4fa;
}
QComboBox::drop-down {
    border: none;
    width: 24px;
}
QComboBox QAbstractItemView {
    background-color: #313244;
    color: #cdd6f4;
    border: 1px solid #45475a;
    selection-background-color: #89b4fa;
    selection-color: #1e1e2e;
}
QTableWidget {
    background-color: #181825;
    color: #cdd6f4;
    border: 1px solid #313244;
    border-radius: 8px;
    gridline-color: #313244;
    alternate-background-color: #1e1e2e;
}
QTableWidget::item {
    padding: 6px 10px;
}
QTableWidget::item:selected {
    background-color: #313244;
    color: #89b4fa;
}
QHeaderView::section {
    background-color: #313244;
    color: #89b4fa;
    padding: 8px 10px;
    border: none;
    border-bottom: 2px solid #45475a;
    font-weight: 600;
}
QTextEdit {
    background-color: #181825;
    color: #a6adc8;
    border: 1px solid #313244;
    border-radius: 8px;
    padding: 8px;
    font-family: 'Cascadia Code', 'Consolas', monospace;
    font-size: 12px;
}
QProgressBar {
    background-color: #313244;
    border: none;
    border-radius: 6px;
    height: 22px;
    text-align: center;
    color: #cdd6f4;
    font-weight: 600;
}
QProgressBar::chunk {
    background-color: #89b4fa;
    border-radius: 6px;
}
QLabel#statusLabel {
    font-size: 14px;
    font-weight: 700;
    padding: 4px 0;
}
QLabel#headerLabel {
    font-size: 18px;
    font-weight: 700;
    color: #89b4fa;
    padding: 8px 0;
}
QLabel#infoLabel {
    color: #6c7086;
    font-size: 12px;
}
QStatusBar {
    background-color: #181825;
    color: #6c7086;
    border-top: 1px solid #313244;
}
QStatusBar::item {
    border: none;
}
QScrollBar:vertical {
    background: #181825;
    width: 10px;
    border: none;
}
QScrollBar::handle:vertical {
    background: #45475a;
    border-radius: 5px;
    min-height: 30px;
}
QScrollBar::handle:vertical:hover {
    background: #585b70;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0;
}
QScrollBar:horizontal {
    background: #181825;
    height: 10px;
    border: none;
}
QScrollBar::handle:horizontal {
    background: #45475a;
    border-radius: 5px;
    min-width: 30px;
}
QScrollBar::handle:horizontal:hover {
    background: #585b70;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    width: 0;
}
QMessageBox {
    background-color: #1e1e2e;
}
QMessageBox QLabel {
    color: #cdd6f4;
    font-size: 14px;
}
QFileDialog {
    background-color: #1e1e2e;
    color: #cdd6f4;
}
QListWidget {
    background-color: #181825;
    color: #cdd6f4;
    border: 1px solid #313244;
    border-radius: 8px;
    padding: 4px;
}
QListWidget::item {
    padding: 4px 8px;
}
QListWidget::item:selected {
    background-color: #313244;
    color: #89b4fa;
}
"""


class SignalBridge(QObject):
    log_signal = pyqtSignal(str, str)
    modem_signal = pyqtSignal()
    progress_signal = pyqtSignal(str)
    smpp_signal = pyqtSignal(str, object)
    call_result_signal = pyqtSignal(dict)
    sms_result_signal = pyqtSignal(dict)
    scan_signal = pyqtSignal(list)
    call_error_signal = pyqtSignal(str)
    sms_error_signal = pyqtSignal(str)
    call_calling_signal = pyqtSignal(str)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("SMPP Gateway v.1.4  -  USB Modem SMS & Voice")
        self.resize(1100, 862)

        # Set application icon
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'app_icon.ico')
        if not os.path.exists(icon_path):
            icon_path = os.path.join(sys._MEIPASS, 'gui', 'app_icon.ico') if getattr(sys, 'frozen', False) else icon_path
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        self.setMinimumSize(900, 690)

        self.modem_manager = ModemManager()
        self.database = Database()
        self.smpp_server = None
        self.tunnel = None
        self.call_handler = CallHandler(self.modem_manager, self.database, callback=self._call_callback)
        self.sms_handler = SmsHandler(self.modem_manager, self.database, callback=self._sms_callback)

        self.call_numbers = []
        self.call_audio_path = None
        self.sms_numbers = []
        self.call_results = []
        self.sms_results = []
        self.call_success_seconds = []   # прослушанные секунды по ответившим (для средней)
        self.call_total = 0              # сколько номеров изначально загружено
        self._number_items = {}          # номер -> QTableWidgetItem (для мгновенного поиска строки
                                          # на больших базах — без него каждый результат звонка
                                          # требовал полного прохода по таблице, O(n²) на 120к номеров)
        self._auto_recall_attempts = 0   # счётчик авто-перезвонов (защита от бесконечного цикла)
        self._call_schedule_fired = False
        self._sms_schedule_fired = False
        self._manual_stop = False        # True, если обзвон остановлен пользователем (не авто-перезванивать)

        self.bridge = SignalBridge()
        self.bridge.log_signal.connect(self._append_log)
        self.bridge.modem_signal.connect(self._refresh_modem_table)
        self.bridge.progress_signal.connect(self._update_progress)
        self.bridge.smpp_signal.connect(self._smpp_status)
        self.bridge.call_result_signal.connect(self._on_call_result)
        self.bridge.call_calling_signal.connect(lambda n: self._set_number_status(n, "Звоним…"))
        self.bridge.sms_result_signal.connect(self._on_sms_result)
        self.bridge.scan_signal.connect(self._on_scan_complete)
        self.bridge.call_error_signal.connect(self._on_call_error)
        self.bridge.sms_error_signal.connect(self._on_sms_error)

        self._init_ui()
        self._load_settings()
        self._auto_scan_and_connect()

        self.timer = QTimer()
        self.timer.timeout.connect(self._sync_and_refresh_modems)
        self.timer.timeout.connect(self._check_schedules)
        self.timer.start(6000)

        self._update_status_bar()

    def _init_ui(self):
        central = QWidget()
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        header = QFrame()
        header.setFixedHeight(56)
        header.setStyleSheet("background-color: #181825; border-bottom: 1px solid #313244;")
        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(20, 0, 20, 0)
        title_label = QLabel("SMPP Gateway")
        title_label.setStyleSheet("font-size: 20px; font-weight: 700; color: #89b4fa; background: transparent;")
        version_label = QLabel("v.1.4")
        version_label.setStyleSheet("font-size: 12px; font-weight: 700; color: #a6e3a1; background: #313244; padding: 2px 8px; border-radius: 4px; margin-left: 6px;")
        subtitle_label = QLabel("USB Modem SMS & Voice Server")
        subtitle_label.setStyleSheet("font-size: 12px; color: #6c7086; background: transparent; margin-left: 8px;")
        header_layout.addWidget(title_label)
        header_layout.addWidget(version_label)
        header_layout.addWidget(subtitle_label)
        header_layout.addStretch()
        header.setLayout(header_layout)
        main_layout.addWidget(header)

        tabs = QTabWidget()
        tabs.addTab(self._modems_tab(), "Модемы")
        tabs.addTab(self._smpp_tab(), "SMPP сервер")
        tabs.addTab(self._calls_tab(), "Массовые звонки")
        tabs.addTab(self._sms_tab(), "SMS рассылка")
        tabs.addTab(self._log_tab(), "Логи")
        tabs.tabBar().setElideMode(Qt.ElideNone)
        tabs.tabBar().setExpanding(False)
        main_layout.addWidget(tabs, 1)

        central.setLayout(main_layout)
        self.setCentralWidget(central)

        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

    # ===================== MODEMS TAB =====================

    def _modems_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(12)

        header = QLabel("Управление модемами")
        header.setObjectName("headerLabel")
        layout.addWidget(header)

        info = QLabel("Программа автоматически находит подключённые USB-модемы. Отметьте галочками те, которые должны работать. E173 поддерживает голосовые вызовы, E3372 - только SMS.")
        info.setObjectName("infoLabel")
        info.setWordWrap(True)
        layout.addWidget(info)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        btn_scan = QPushButton("Сканировать порты")
        btn_scan.clicked.connect(self._auto_scan_and_connect)
        btn_connect_all = QPushButton("Подключить все")
        btn_connect_all.clicked.connect(self._connect_all)
        btn_disconnect_all = QPushButton("Отключить все")
        btn_disconnect_all.clicked.connect(self._disconnect_all)
        btn_row.addWidget(btn_scan)
        btn_row.addWidget(btn_connect_all)
        btn_row.addWidget(btn_disconnect_all)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        self.modem_table = QTableWidget(0, 7)
        self.modem_table.setHorizontalHeaderLabels(["Вкл", "Порт", "Модель", "IMEI", "Голос", "Сигнал", "Статус"])
        self.modem_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.modem_table.horizontalHeader().setStretchLastSection(True)
        self.modem_table.setAlternatingRowColors(True)
        self.modem_table.verticalHeader().setVisible(False)
        self.modem_table.setShowGrid(False)
        self.modem_table.setWordWrap(True)
        self.modem_table.setTextElideMode(Qt.ElideNone)
        self.modem_table.setColumnWidth(0, 50)
        self.modem_table.setColumnWidth(1, 100)
        self.modem_table.setColumnWidth(2, 120)
        self.modem_table.setColumnWidth(3, 160)
        self.modem_table.setColumnWidth(4, 80)
        self.modem_table.setColumnWidth(5, 80)
        self.modem_table.setColumnWidth(6, 120)
        layout.addWidget(self.modem_table)

        log_label = QLabel("Лог модемов")
        log_label.setStyleSheet("font-weight: 600; color: #a6adc8; margin-top: 4px;")
        layout.addWidget(log_label)
        self.modem_log = QTextEdit()
        self.modem_log.setReadOnly(True)
        self.modem_log.setMinimumHeight(240)
        layout.addWidget(self.modem_log, 1)

        widget.setLayout(layout)
        return widget

    def _auto_scan_and_connect(self):
        threading.Thread(target=self._scan_connect_thread, daemon=True).start()

    def _scan_connect_thread(self):
        self.bridge.log_signal.emit("modem", "Сканирование портов...")
        groups = self.modem_manager.scan_all_detailed()
        self.bridge.log_signal.emit("modem", f"Найдено модемов (по USB): {len(groups)}")
        for g in groups:
            self.bridge.log_signal.emit(
                "modem",
                f"  модем: команд.порт={g.get('at_port')}, голос.порт={g.get('voice_port')}, modem={g.get('modem_port')}"
            )
        results = self.modem_manager.auto_connect_all()
        if not results:
            self.bridge.log_signal.emit("modem", "Модемы не найдены. Проверьте подключение USB-модемов и драйвер Huawei.")
        else:
            ok_count = sum(1 for r in results if r['success'])
            for r in results:
                if r['success']:
                    voice = "да" if r['voice'] else "нет"
                    self.bridge.log_signal.emit("modem", f"✓ Подключён {r['port']}: {r['model']}, голос: {voice}")
                else:
                    reason = r.get('reason', 'нет ответа')
                    self.bridge.log_signal.emit("modem", f"✗ Не удалось: {r['port']} ({reason})")
            self.bridge.log_signal.emit("modem", f"Итого подключено: {ok_count} из {len(results)}")
        self.bridge.modem_signal.emit()

    def _connect_all(self):
        threading.Thread(target=self._scan_connect_thread, daemon=True).start()

    def _disconnect_all(self):
        self.modem_manager.disconnect_all()
        self._append_log("modem", "Все модемы отключены")
        self._refresh_modem_table()

    def _sync_and_refresh_modems(self):
        removed = self.modem_manager.sync_modems()
        if removed:
            for port in removed:
                self._append_log("modem", f"Модем отключён: {port}")
        self._refresh_modem_table()

    def _on_call_schedule_toggled(self, on):
        if on:
            self._call_schedule_fired = False
            self.call_schedule_status.setText("")

    def _on_sms_schedule_toggled(self, on):
        if on:
            self._sms_schedule_fired = False
            self.sms_schedule_status.setText("")

    def _check_schedules(self):
        """Проверяет запланированное время запуска обзвона/рассылки (пункт «Авто-отправка по расписанию»)."""
        now = QDateTime.currentDateTime()
        if (self.call_schedule_enabled.isChecked() and not self._call_schedule_fired
                and now >= self.call_schedule_time.dateTime()):
            self._call_schedule_fired = True
            self.call_schedule_enabled.setChecked(False)
            self.call_schedule_status.setText(f"Запущено автоматически в {now.toString('HH:mm:ss')}")
            self._append_log("event", "Автозапуск обзвона по расписанию")
            self._start_calls()

        if (self.sms_schedule_enabled.isChecked() and not self._sms_schedule_fired
                and now >= self.sms_schedule_time.dateTime()):
            self._sms_schedule_fired = True
            self.sms_schedule_enabled.setChecked(False)
            self.sms_schedule_status.setText(f"Запущено автоматически в {now.toString('HH:mm:ss')}")
            self._append_log("event", "Автозапуск рассылки по расписанию")
            self._start_sms_campaign()

    def _fill_modem_row(self, row, s, with_checkbox):
        if with_checkbox:
            chk = QCheckBox()
            chk.setChecked(s.get('active', True))
            chk.setStyleSheet("QCheckBox { margin-left: 12px; }")
            chk.stateChanged.connect(lambda state, p=s['port']: self._on_modem_checkbox(p, state))
            self.modem_table.setCellWidget(row, 0, chk)
        self.modem_table.setItem(row, 1, QTableWidgetItem(s['port']))
        self.modem_table.setItem(row, 2, QTableWidgetItem(s['model'] or "?"))
        self.modem_table.setItem(row, 3, QTableWidgetItem(s['imei'] or "?"))
        voice_item = QTableWidgetItem("Да" if s['supports_voice'] else "Нет")
        voice_item.setForeground(QColor("#a6e3a1") if s['supports_voice'] else QColor("#f38ba8"))
        self.modem_table.setItem(row, 4, voice_item)
        sig = str(s['signal_quality']) if s['signal_quality'] else "?"
        sig_item = QTableWidgetItem(sig)
        if s['signal_quality'] and s['signal_quality'] > 15:
            sig_item.setForeground(QColor("#a6e3a1"))
        elif s['signal_quality'] and s['signal_quality'] > 5:
            sig_item.setForeground(QColor("#f9e2af"))
        elif s['signal_quality']:
            sig_item.setForeground(QColor("#f38ba8"))
        self.modem_table.setItem(row, 5, sig_item)
        status_item = QTableWidgetItem("Подключён" if s['connected'] else "Отключён")
        status_item.setForeground(QColor("#a6e3a1") if s['connected'] else QColor("#f38ba8"))
        self.modem_table.setItem(row, 6, status_item)

    def _refresh_modem_table(self):
        statuses = self.modem_manager.get_all_status()
        # если состав портов не изменился — обновляем ячейки на месте (скролл не прыгает)
        same = (self.modem_table.rowCount() == len(statuses) and all(
            self.modem_table.item(r, 1) and self.modem_table.item(r, 1).text() == statuses[r]['port']
            for r in range(len(statuses))))
        if same:
            for r, s in enumerate(statuses):
                self._fill_modem_row(r, s, with_checkbox=False)
        else:
            self.modem_table.setRowCount(0)
            for s in statuses:
                row = self.modem_table.rowCount()
                self.modem_table.insertRow(row)
                self._fill_modem_row(row, s, with_checkbox=True)
        self._update_status_bar()

    def _on_modem_checkbox(self, port, state):
        active = state == Qt.Checked
        self.modem_manager.set_active(port, active)
        self._update_status_bar()

    # ===================== SMPP TAB =====================

    def _smpp_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(8)

        header = QLabel("SMPP сервер")
        header.setObjectName("headerLabel")
        layout.addWidget(header)

        info = QLabel("SMPP 3.4 сервер принимает запросы от удалённого SMPP-клиента и отправляет SMS через модемы.")
        info.setObjectName("infoLabel")
        info.setWordWrap(True)
        layout.addWidget(info)

        # ===== Две колонки: слева настройки, справа лог =====
        columns = QHBoxLayout()
        columns.setSpacing(14)

        # ---- ЛЕВАЯ КОЛОНКА: настройки ----
        left = QVBoxLayout()
        left.setSpacing(10)

        FW = 170  # ширина полей в компактной колонке

        ip_group = QGroupBox("Адрес сервера")
        ip_layout = QFormLayout()
        ip_layout.setSpacing(8)
        ip_layout.setContentsMargins(12, 14, 12, 12)
        ip_layout.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        self.local_ip_label = QLabel(get_local_ip())
        self.local_ip_label.setStyleSheet("font-weight: 600; color: #a6e3a1;")
        self.smpp_domain = QLineEdit()
        self.smpp_domain.setPlaceholderText("your-domain.example")
        self.smpp_domain.setMinimumHeight(30)
        self.smpp_domain.setMinimumWidth(FW)
        ip_layout.addRow("Локальный IP:", self.local_ip_label)
        ip_layout.addRow("Домен:", self.smpp_domain)
        ip_group.setLayout(ip_layout)
        left.addWidget(ip_group)

        settings_group = QGroupBox("Настройки подключения")
        form = QFormLayout()
        form.setSpacing(10)
        form.setContentsMargins(12, 14, 12, 12)
        form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        self.smpp_host = QLineEdit("0.0.0.0")
        self.smpp_port = QSpinBox()
        self.smpp_port.setRange(1, 65535)
        self.smpp_port.setValue(2775)
        self.smpp_port.valueChanged.connect(self._update_smpp_conn_info)
        self.smpp_system_id = QLineEdit("smppuser")
        self.smpp_password = QLineEdit("smpppass")
        self.smpp_password.setEchoMode(QLineEdit.Password)
        self.smpp_system_type = QLineEdit("SMPP")
        for w in (self.smpp_host, self.smpp_port, self.smpp_system_id,
                  self.smpp_password, self.smpp_system_type):
            w.setMinimumHeight(30)
            w.setMinimumWidth(FW)
        form.addRow("Хост:", self.smpp_host)
        form.addRow("Порт:", self.smpp_port)
        form.addRow("System ID:", self.smpp_system_id)
        form.addRow("Пароль:", self.smpp_password)
        form.addRow("System Type:", self.smpp_system_type)
        settings_group.setLayout(form)
        left.addWidget(settings_group)

        pub_group = QGroupBox("Доступ извне (через сервер)")
        pub_layout = QVBoxLayout()
        pub_layout.setContentsMargins(12, 14, 12, 12)
        pub_layout.setSpacing(6)
        self.smpp_tunnel_enabled = QCheckBox("Сделать доступным из интернета (туннель)")
        pub_layout.addWidget(self.smpp_tunnel_enabled)
        self.smpp_tunnel_status = QLabel("Туннель: выключен")
        self.smpp_tunnel_status.setStyleSheet("color: #6c7086; font-weight: 600;")
        pub_layout.addWidget(self.smpp_tunnel_status)
        pub_hint = QLabel("Внешний клиент подключается к домену с System ID и паролем. "
                          "Работает без белого IP и проброса портов.")
        pub_hint.setStyleSheet("color: #6c7086; font-size: 11px;")
        pub_hint.setWordWrap(True)
        pub_layout.addWidget(pub_hint)
        pub_group.setLayout(pub_layout)
        left.addWidget(pub_group)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        self.btn_smpp_start = QPushButton("Запустить сервер")
        self.btn_smpp_start.setObjectName("smppStartBtn")
        self.btn_smpp_start.clicked.connect(self._start_smpp)
        self.btn_smpp_stop = QPushButton("Остановить")
        self.btn_smpp_stop.setObjectName("smppStopBtn")
        self.btn_smpp_stop.clicked.connect(self._stop_smpp)
        self.btn_smpp_stop.setEnabled(False)
        btn_row.addWidget(self.btn_smpp_start)
        btn_row.addWidget(self.btn_smpp_stop)
        left.addLayout(btn_row)

        self.smpp_status_label = QLabel("Сервер остановлен")
        self.smpp_status_label.setObjectName("statusLabel")
        self.smpp_status_label.setStyleSheet("color: #f38ba8; font-weight: 700; font-size: 14px;")
        left.addWidget(self.smpp_status_label)

        conn_info = QLabel(f"Клиент подключается к: {get_local_ip()}:2775")
        conn_info.setStyleSheet("color: #89b4fa; font-size: 12px; font-weight: 600;")
        conn_info.setWordWrap(True)
        self.smpp_conn_info = conn_info
        left.addWidget(conn_info)
        self.smpp_domain.textChanged.connect(self._update_smpp_conn_info)
        left.addStretch()

        left_widget = QWidget()
        left_widget.setLayout(left)
        left_widget.setFixedWidth(360)
        columns.addWidget(left_widget)

        # ---- ПРАВАЯ КОЛОНКА: лог (занимает всё оставшееся место) ----
        right = QVBoxLayout()
        right.setSpacing(6)
        log_label = QLabel("Лог SMPP")
        log_label.setStyleSheet("font-weight: 600; color: #a6adc8;")
        right.addWidget(log_label)
        self.smpp_log = QTextEdit()
        self.smpp_log.setReadOnly(True)
        right.addWidget(self.smpp_log, 1)
        columns.addLayout(right, 1)

        layout.addLayout(columns, 1)
        widget.setLayout(layout)
        return widget

    # ===================== CALLS TAB =====================

    def _calls_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(12)

        header = QLabel("Массовые звонки")
        header.setObjectName("headerLabel")
        layout.addWidget(header)

        info = QLabel("Загрузите базу номеров и аудиофайл (WAV или MP3). Звонки выполняются через модемы E173 с поддержкой голоса.")
        info.setObjectName("infoLabel")
        info.setWordWrap(True)
        layout.addWidget(info)

        file_group = QGroupBox("Файлы")
        file_layout = QVBoxLayout()
        file_layout.setSpacing(10)

        numbers_row = QHBoxLayout()
        btn_load_numbers = QPushButton("Загрузить номера (CSV/TXT/Excel)")
        btn_load_numbers.clicked.connect(self._load_call_numbers)
        self.call_numbers_label = QLabel("База номеров не загружена")
        self.call_numbers_label.setStyleSheet("color: #6c7086; font-style: italic;")
        numbers_row.addWidget(btn_load_numbers)
        numbers_row.addWidget(self.call_numbers_label)
        numbers_row.addStretch()
        file_layout.addLayout(numbers_row)

        audio_row = QHBoxLayout()
        btn_load_audio = QPushButton("Выбрать аудио (WAV / MP3)")
        btn_load_audio.clicked.connect(self._load_call_audio)
        self.call_audio_label = QLabel("Аудиофайл не выбран")
        self.call_audio_label.setStyleSheet("color: #6c7086; font-style: italic;")
        audio_row.addWidget(btn_load_audio)
        audio_row.addWidget(self.call_audio_label)
        audio_row.addStretch()
        file_layout.addLayout(audio_row)

        # Таймаут ожидания
        timeout_row = QHBoxLayout()
        timeout_row.setSpacing(10)
        timeout_label = QLabel("Ждать ответа (сек):")
        timeout_label.setStyleSheet("font-weight: 500; color: #a6adc8; font-size: 12px;")
        timeout_label.setToolTip("Сколько ждать, пока абонент снимет трубку. Меньше 15 сек ставить нельзя — "
                                 "звонок не успеет дойти до абонента.")
        timeout_row.addWidget(timeout_label)
        self.call_timeout = QSpinBox()
        self.call_timeout.setRange(15, 120)
        self.call_timeout.setValue(30)
        self.call_timeout.setMaximumWidth(80)
        timeout_row.addWidget(self.call_timeout)
        timeout_row.addStretch()
        file_layout.addLayout(timeout_row)

        # Кнопки обзвона — под таймаутом
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        self.btn_start_calls = QPushButton("Начать обзвон")
        self.btn_start_calls.setObjectName("startBtn")
        self.btn_start_calls.clicked.connect(self._start_calls)
        self.btn_stop_calls = QPushButton("Остановить")
        self.btn_stop_calls.setObjectName("stopBtn")
        self.btn_stop_calls.clicked.connect(self._stop_calls)
        self.btn_stop_calls.setEnabled(False)
        btn_row.addWidget(self.btn_start_calls)
        btn_row.addWidget(self.btn_stop_calls)
        btn_row.addStretch()
        file_layout.addLayout(btn_row)

        file_group.setLayout(file_layout)

        # --- Авто-обзвон и расписание ---
        auto_group = QGroupBox("Автоматизация")
        auto_layout = QVBoxLayout()
        auto_layout.setContentsMargins(12, 12, 12, 12)
        auto_layout.setSpacing(8)

        self.call_auto_recall = QCheckBox("Автоматически перезванивать неотвеченным")
        self.call_auto_recall.setToolTip(
            "После завершения обзвона программа сама через минуту повторит попытку "
            "для номеров со статусом «Нет ответа»/«Сброшен»/«Занято» (до 3 раз).")
        auto_layout.addWidget(self.call_auto_recall)

        sched_row = QHBoxLayout()
        sched_row.setSpacing(8)
        self.call_schedule_enabled = QCheckBox("Запустить обзвон автоматически в:")
        self.call_schedule_enabled.toggled.connect(self._on_call_schedule_toggled)
        self.call_schedule_time = QDateTimeEdit(QDateTime.currentDateTime().addSecs(600))
        self.call_schedule_time.setDisplayFormat("dd.MM.yyyy HH:mm")
        self.call_schedule_time.setCalendarPopup(True)
        self.call_schedule_time.setMaximumWidth(160)
        sched_row.addWidget(self.call_schedule_enabled)
        sched_row.addWidget(self.call_schedule_time)
        sched_row.addStretch()
        auto_layout.addLayout(sched_row)
        self.call_schedule_status = QLabel("")
        self.call_schedule_status.setStyleSheet("color: #6c7086; font-size: 11px;")
        auto_layout.addWidget(self.call_schedule_status)

        auto_group.setLayout(auto_layout)

        # ===== Две колонки: слева файлы+статистика, справа таблицы =====
        columns = QHBoxLayout()
        columns.setSpacing(14)
        left = QVBoxLayout()
        left.setSpacing(10)
        left.addWidget(file_group)
        left.addWidget(auto_group)

        self.call_progress = QProgressBar()
        self.call_progress.setVisible(False)
        left.addWidget(self.call_progress)

        self.call_status_label = QLabel("")
        self.call_status_label.setStyleSheet("color: #a6adc8; font-weight: 500;")
        left.addWidget(self.call_status_label)

        # --- Панель статистики кампании (сетка 2x2 для узкой колонки) ---
        stats_group = QGroupBox("Статистика обзвона")
        stats_row = QGridLayout()
        stats_row.setContentsMargins(12, 12, 12, 12)
        stats_row.setHorizontalSpacing(24)
        stats_row.setVerticalSpacing(10)

        def _stat_label():
            l = QLabel("—")
            l.setStyleSheet("font-size: 15px; font-weight: 700; color: #89b4fa;")
            return l

        def _stat_caption(text):
            c = QLabel(text)
            c.setStyleSheet("color: #6c7086; font-size: 11px;")
            return c

        col1 = QVBoxLayout()
        self.stat_processed = _stat_label()
        col1.addWidget(self.stat_processed)
        col1.addWidget(_stat_caption("Обработано"))

        col2 = QVBoxLayout()
        self.stat_answered = _stat_label()
        col2.addWidget(self.stat_answered)
        col2.addWidget(_stat_caption("Ответили"))

        col3 = QVBoxLayout()
        self.stat_avg = _stat_label()
        col3.addWidget(self.stat_avg)
        col3.addWidget(_stat_caption("Средняя прослушано"))

        col4 = QVBoxLayout()
        self.stat_percent = _stat_label()
        col4.addWidget(self.stat_percent)
        col4.addWidget(_stat_caption("Процент обработки"))

        stats_row.addLayout(col1, 0, 0)
        stats_row.addLayout(col2, 0, 1)
        stats_row.addLayout(col3, 1, 0)
        stats_row.addLayout(col4, 1, 1)
        stats_row.setColumnStretch(2, 1)
        stats_group.setLayout(stats_row)
        left.addWidget(stats_group)
        left.addStretch()
        left_widget = QWidget()
        left_widget.setLayout(left)
        left_widget.setFixedWidth(380)
        columns.addWidget(left_widget)
        self._reset_call_stats()

        # ---- ПРАВАЯ КОЛОНКА: таблицы Номера/Успешные ----
        sub_tabs = QTabWidget()

        numbers_tab = QWidget()
        numbers_layout = QVBoxLayout()
        numbers_layout.setSpacing(8)
        num_btn_row = QHBoxLayout()
        btn_num_all = QPushButton("Выбрать все")
        btn_num_all.clicked.connect(lambda: self._set_all_checks(self.call_numbers_list, True))
        btn_num_none = QPushButton("Снять все")
        btn_num_none.clicked.connect(lambda: self._set_all_checks(self.call_numbers_list, False))
        btn_num_del = QPushButton("Удалить выбранные")
        btn_num_del.clicked.connect(lambda: self._delete_checked(self.call_numbers_list, is_numbers=True))
        btn_num_export = QPushButton("Экспорт оставшихся")
        btn_num_export.setToolTip("Сохранить в файл номера, оставшиеся в списке (например после остановки обзвона).")
        btn_num_export.clicked.connect(self._export_remaining_numbers)
        num_btn_row.addWidget(btn_num_all)
        num_btn_row.addWidget(btn_num_none)
        num_btn_row.addWidget(btn_num_del)
        num_btn_row.addWidget(btn_num_export)
        num_btn_row.addStretch()
        btn_num_recall = QPushButton("Обзвонить неотвеченные")
        btn_num_recall.setToolTip("Позвонить всем номерам, кроме уже ответивших "
                                  "(в очереди, нет ответа, сброшен, занято).")
        btn_num_recall.clicked.connect(self._call_unanswered)
        num_btn_row.addWidget(btn_num_recall)
        numbers_layout.addLayout(num_btn_row)
        self.call_numbers_list = QTableWidget(0, 3)
        self.call_numbers_list.setHorizontalHeaderLabels(["✓", "Номер", "Статус"])
        self.call_numbers_list.horizontalHeader().setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.call_numbers_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.call_numbers_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
        self.call_numbers_list.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.call_numbers_list.setAlternatingRowColors(True)
        self.call_numbers_list.verticalHeader().setVisible(False)
        self.call_numbers_list.setShowGrid(False)
        self.call_numbers_list.setWordWrap(False)
        self.call_numbers_list.setTextElideMode(Qt.ElideNone)
        self.call_numbers_list.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.call_numbers_list.setColumnWidth(0, 44)
        self.call_numbers_list.setColumnWidth(1, 200)
        numbers_layout.addWidget(self.call_numbers_list)
        numbers_tab.setLayout(numbers_layout)
        sub_tabs.addTab(numbers_tab, "Номера")

        success_tab = QWidget()
        success_layout = QVBoxLayout()
        success_layout.setSpacing(8)
        succ_btn_row = QHBoxLayout()
        btn_succ_all = QPushButton("Выбрать все")
        btn_succ_all.clicked.connect(lambda: self._set_all_checks(self.call_success_table, True))
        btn_succ_none = QPushButton("Снять все")
        btn_succ_none.clicked.connect(lambda: self._set_all_checks(self.call_success_table, False))
        btn_succ_del = QPushButton("Удалить выбранные")
        btn_succ_del.clicked.connect(lambda: self._delete_checked(self.call_success_table))
        succ_btn_row.addWidget(btn_succ_all)
        succ_btn_row.addWidget(btn_succ_none)
        succ_btn_row.addWidget(btn_succ_del)
        succ_btn_row.addStretch()
        success_layout.addLayout(succ_btn_row)
        self.call_success_table = QTableWidget(0, 6)
        self.call_success_table.setHorizontalHeaderLabels(["✓", "Номер", "Модем", "Всего (сек)", "Процент", "Время звонка"])
        self.call_success_table.horizontalHeader().setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.call_success_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        for _c in range(1, 6):
            self.call_success_table.horizontalHeader().setSectionResizeMode(_c, QHeaderView.Interactive)
        self.call_success_table.horizontalHeader().setStretchLastSection(True)
        self.call_success_table.setAlternatingRowColors(True)
        self.call_success_table.verticalHeader().setVisible(False)
        self.call_success_table.setShowGrid(False)
        self.call_success_table.setWordWrap(True)
        self.call_success_table.setTextElideMode(Qt.ElideNone)
        self.call_success_table.setColumnWidth(0, 44)
        self.call_success_table.setColumnWidth(1, 170)
        self.call_success_table.setColumnWidth(2, 110)
        self.call_success_table.setColumnWidth(3, 110)
        self.call_success_table.setColumnWidth(4, 90)
        self.call_success_table.setColumnWidth(5, 110)
        success_layout.addWidget(self.call_success_table)
        success_tab.setLayout(success_layout)
        sub_tabs.addTab(success_tab, "Успешные")

        self.call_sub_tabs = sub_tabs
        sub_tabs.tabBar().setElideMode(Qt.ElideNone)
        sub_tabs.tabBar().setExpanding(False)
        columns.addWidget(sub_tabs, 1)

        layout.addLayout(columns, 1)
        widget.setLayout(layout)
        return widget

    # ===================== SMS TAB =====================

    def _sms_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(12)

        header = QLabel("SMS рассылка")
        header.setObjectName("headerLabel")
        layout.addWidget(header)

        info = QLabel("Загрузите базу номеров, введите текст сообщения и начните рассылку. SMS отправляются через все активные модемы.")
        info.setObjectName("infoLabel")
        info.setWordWrap(True)
        layout.addWidget(info)

        file_group = QGroupBox("База номеров")
        file_layout = QHBoxLayout()
        file_layout.setContentsMargins(12, 12, 12, 12)
        self.sms_numbers_label = QLabel("База номеров не загружена")
        self.sms_numbers_label.setStyleSheet("color: #6c7086; font-style: italic;")
        btn_load_sms_numbers = QPushButton("Загрузить (CSV/TXT/Excel)")
        btn_load_sms_numbers.clicked.connect(self._load_sms_numbers)
        file_layout.addWidget(btn_load_sms_numbers)
        file_layout.addWidget(self.sms_numbers_label)
        file_layout.addStretch()
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)

        msg_group = QGroupBox("Текст сообщения")
        msg_layout = QVBoxLayout()
        msg_layout.setSpacing(6)
        msg_layout.setContentsMargins(12, 12, 12, 12)
        self.sms_message = QTextEdit()
        self.sms_message.setPlaceholderText("Введите текст SMS...")
        self.sms_message.setMaximumHeight(100)
        self.sms_message.setStyleSheet("QTextEdit { font-size: 14px; }")
        msg_layout.addWidget(self.sms_message)
        counter_row = QHBoxLayout()
        self.sms_counter = QLabel("0 символов, 1 SMS")
        self.sms_counter.setStyleSheet("color: #6c7086; font-size: 12px;")
        self.sms_message.textChanged.connect(self._update_sms_counter)
        counter_row.addWidget(self.sms_counter)
        counter_row.addStretch()
        msg_layout.addLayout(counter_row)
        msg_group.setLayout(msg_layout)
        layout.addWidget(msg_group)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        self.btn_start_sms = QPushButton("Начать рассылку")
        self.btn_start_sms.setObjectName("startBtn")
        self.btn_start_sms.clicked.connect(self._start_sms_campaign)
        self.btn_stop_sms = QPushButton("Остановить")
        self.btn_stop_sms.setObjectName("stopBtn")
        self.btn_stop_sms.clicked.connect(self._stop_sms_campaign)
        self.btn_stop_sms.setEnabled(False)
        btn_row.addWidget(self.btn_start_sms)
        btn_row.addWidget(self.btn_stop_sms)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        sched_group = QGroupBox("Автоматизация")
        sched_layout = QHBoxLayout()
        sched_layout.setContentsMargins(12, 10, 12, 10)
        sched_layout.setSpacing(8)
        self.sms_schedule_enabled = QCheckBox("Запустить рассылку автоматически в:")
        self.sms_schedule_enabled.toggled.connect(self._on_sms_schedule_toggled)
        self.sms_schedule_time = QDateTimeEdit(QDateTime.currentDateTime().addSecs(600))
        self.sms_schedule_time.setDisplayFormat("dd.MM.yyyy HH:mm")
        self.sms_schedule_time.setCalendarPopup(True)
        self.sms_schedule_time.setMaximumWidth(160)
        sched_layout.addWidget(self.sms_schedule_enabled)
        sched_layout.addWidget(self.sms_schedule_time)
        sched_layout.addStretch()
        sched_group.setLayout(sched_layout)
        layout.addWidget(sched_group)
        self.sms_schedule_status = QLabel("")
        self.sms_schedule_status.setStyleSheet("color: #6c7086; font-size: 11px;")
        layout.addWidget(self.sms_schedule_status)

        self.sms_progress = QProgressBar()
        self.sms_progress.setVisible(False)
        layout.addWidget(self.sms_progress)

        self.sms_status_label = QLabel("")
        self.sms_status_label.setStyleSheet("color: #a6adc8; font-weight: 500;")
        layout.addWidget(self.sms_status_label)

        sub_tabs = QTabWidget()

        numbers_tab = QWidget()
        numbers_layout = QVBoxLayout()
        numbers_layout.setSpacing(8)
        self.sms_numbers_list = QTableWidget(0, 1)
        self.sms_numbers_list.setHorizontalHeaderLabels(["Номер"])
        self.sms_numbers_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.sms_numbers_list.setAlternatingRowColors(True)
        self.sms_numbers_list.verticalHeader().setVisible(False)
        self.sms_numbers_list.setShowGrid(False)
        self.sms_numbers_list.setWordWrap(True)
        self.sms_numbers_list.setTextElideMode(Qt.ElideNone)
        self.sms_numbers_list.setColumnWidth(0, 400)
        numbers_layout.addWidget(self.sms_numbers_list)
        numbers_tab.setLayout(numbers_layout)
        sub_tabs.addTab(numbers_tab, "Номера")

        success_tab = QWidget()
        success_layout = QVBoxLayout()
        success_layout.setSpacing(8)
        self.sms_success_table = QTableWidget(0, 3)
        self.sms_success_table.setHorizontalHeaderLabels(["Номер", "Модем", "Статус"])
        self.sms_success_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Interactive)
        self.sms_success_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Interactive)
        self.sms_success_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Interactive)
        self.sms_success_table.horizontalHeader().setStretchLastSection(True)
        self.sms_success_table.setAlternatingRowColors(True)
        self.sms_success_table.verticalHeader().setVisible(False)
        self.sms_success_table.setShowGrid(False)
        self.sms_success_table.setWordWrap(True)
        self.sms_success_table.setTextElideMode(Qt.ElideNone)
        self.sms_success_table.setColumnWidth(0, 200)
        self.sms_success_table.setColumnWidth(1, 120)
        self.sms_success_table.setColumnWidth(2, 200)
        success_layout.addWidget(self.sms_success_table)
        success_tab.setLayout(success_layout)
        sub_tabs.addTab(success_tab, "Успешные")

        failed_tab = QWidget()
        failed_layout = QVBoxLayout()
        failed_layout.setSpacing(8)
        self.sms_failed_table = QTableWidget(0, 3)
        self.sms_failed_table.setHorizontalHeaderLabels(["Номер", "Модем", "Причина"])
        self.sms_failed_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Interactive)
        self.sms_failed_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Interactive)
        self.sms_failed_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Interactive)
        self.sms_failed_table.horizontalHeader().setStretchLastSection(True)
        self.sms_failed_table.setAlternatingRowColors(True)
        self.sms_failed_table.verticalHeader().setVisible(False)
        self.sms_failed_table.setShowGrid(False)
        self.sms_failed_table.setWordWrap(True)
        self.sms_failed_table.setTextElideMode(Qt.ElideNone)
        self.sms_failed_table.setColumnWidth(0, 180)
        self.sms_failed_table.setColumnWidth(1, 120)
        self.sms_failed_table.setColumnWidth(2, 300)
        failed_layout.addWidget(self.sms_failed_table)
        failed_tab.setLayout(failed_layout)
        sub_tabs.addTab(failed_tab, "Неуспешные")

        self.sms_sub_tabs = sub_tabs
        sub_tabs.tabBar().setElideMode(Qt.ElideNone)
        sub_tabs.tabBar().setExpanding(False)
        layout.addWidget(sub_tabs, 1)

        widget.setLayout(layout)
        return widget

    # ===================== LOG TAB =====================

    def _log_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(12)

        header = QLabel("Логи")
        header.setObjectName("headerLabel")
        layout.addWidget(header)

        info = QLabel("Живой лог показывает процесс обзвона и рассылки в реальном времени. Ниже — история из базы данных (кнопка «Обновить»).")
        info.setObjectName("infoLabel")
        info.setWordWrap(True)
        layout.addWidget(info)

        live_label = QLabel("Живой лог событий")
        live_label.setStyleSheet("font-weight: 600; color: #a6adc8;")
        layout.addWidget(live_label)
        live_row = QHBoxLayout()
        self.event_log = QTextEdit()
        self.event_log.setReadOnly(True)
        self.event_log.setMinimumHeight(160)
        layout.addWidget(self.event_log)
        btn_clear_log = QPushButton("Очистить живой лог")
        btn_clear_log.clicked.connect(lambda: self.event_log.clear())
        live_row.addWidget(btn_clear_log)
        live_row.addStretch()
        layout.addLayout(live_row)

        tabs = QTabWidget()

        sms_log_tab = QWidget()
        sms_log_layout = QVBoxLayout()
        sms_log_layout.setSpacing(8)
        self.sms_log_table = QTableWidget(0, 8)
        self.sms_log_table.setHorizontalHeaderLabels(["ID", "Время", "Номер", "Сообщение", "Модем", "Статус", "Источник", "Детали"])
        self.sms_log_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.sms_log_table.horizontalHeader().setStretchLastSection(True)
        self.sms_log_table.setAlternatingRowColors(True)
        self.sms_log_table.verticalHeader().setVisible(False)
        self.sms_log_table.setShowGrid(False)
        self.sms_log_table.setWordWrap(True)
        self.sms_log_table.setTextElideMode(Qt.ElideNone)
        self.sms_log_table.setColumnWidth(0, 50)
        self.sms_log_table.setColumnWidth(1, 160)
        self.sms_log_table.setColumnWidth(2, 140)
        self.sms_log_table.setColumnWidth(3, 200)
        self.sms_log_table.setColumnWidth(4, 100)
        self.sms_log_table.setColumnWidth(5, 100)
        self.sms_log_table.setColumnWidth(6, 100)
        self.sms_log_table.setColumnWidth(7, 200)
        sms_log_layout.addWidget(self.sms_log_table)
        btn_refresh_sms_log = QPushButton("Обновить")
        btn_refresh_sms_log.clicked.connect(self._refresh_sms_log)
        btn_row_sms = QHBoxLayout()
        btn_row_sms.addWidget(btn_refresh_sms_log)
        btn_row_sms.addStretch()
        sms_log_layout.addLayout(btn_row_sms)
        sms_log_tab.setLayout(sms_log_layout)
        tabs.addTab(sms_log_tab, "SMS логи")

        call_log_tab = QWidget()
        call_log_layout = QVBoxLayout()
        call_log_layout.setSpacing(8)
        self.call_log_table = QTableWidget(0, 7)
        self.call_log_table.setHorizontalHeaderLabels(["ID", "Время", "Номер", "Модем", "Статус", "Аудио", "Длительность"])
        self.call_log_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.call_log_table.horizontalHeader().setStretchLastSection(True)
        self.call_log_table.setAlternatingRowColors(True)
        self.call_log_table.verticalHeader().setVisible(False)
        self.call_log_table.setShowGrid(False)
        self.call_log_table.setWordWrap(True)
        self.call_log_table.setTextElideMode(Qt.ElideNone)
        self.call_log_table.setColumnWidth(0, 50)
        self.call_log_table.setColumnWidth(1, 160)
        self.call_log_table.setColumnWidth(2, 140)
        self.call_log_table.setColumnWidth(3, 100)
        self.call_log_table.setColumnWidth(4, 120)
        self.call_log_table.setColumnWidth(5, 200)
        self.call_log_table.setColumnWidth(6, 160)
        call_log_layout.addWidget(self.call_log_table)
        btn_refresh_call_log = QPushButton("Обновить")
        btn_refresh_call_log.clicked.connect(self._refresh_call_log)
        btn_row_call = QHBoxLayout()
        btn_row_call.addWidget(btn_refresh_call_log)
        btn_row_call.addStretch()
        call_log_layout.addLayout(btn_row_call)
        call_log_tab.setLayout(call_log_layout)
        tabs.addTab(call_log_tab, "Логи звонков")

        tabs.tabBar().setElideMode(Qt.ElideNone)
        tabs.tabBar().setExpanding(False)
        layout.addWidget(tabs, 1)
        widget.setLayout(layout)
        return widget

    # ===================== HANDLERS =====================

    def _append_log(self, target, message):
        timestamp = time.strftime("%H:%M:%S")
        line = f"[{timestamp}] {message}"
        if target == "modem":
            self.modem_log.append(line)
            self._log_event(line)
        elif target == "smpp":
            self.smpp_log.append(line)
            self._log_event(line)
        elif target == "event":
            self._log_event(line)

    def _log_event(self, line):
        if hasattr(self, "event_log"):
            self.event_log.append(line)
            # ограничим объём, чтобы не разрасталось
            if self.event_log.document().blockCount() > 500:
                self.event_log.clear()

    def _on_scan_complete(self, results):
        self._refresh_modem_table()

    # --- SMPP ---

    def _update_smpp_conn_info(self):
        domain = self.smpp_domain.text().strip()
        port = self.smpp_port.value()
        if domain:
            self.smpp_conn_info.setText(f"Клиент должен подключаться к: {domain}:{port}")
        else:
            self.smpp_conn_info.setText(f"Клиент должен подключаться к: {get_local_ip()}:{port}")

    def _start_smpp(self):
        host = self.smpp_host.text()
        port = self.smpp_port.value()
        auth = {
            'system_id': self.smpp_system_id.text(),
            'password': self.smpp_password.text(),
        }
        self.smpp_server = SMPPServer(
            host=host, port=port, auth=auth,
            sms_handler=self.sms_handler, database=self.database,
            callback=self._smpp_callback
        )
        self.smpp_server.start()
        self.btn_smpp_start.setEnabled(False)
        self.btn_smpp_stop.setEnabled(True)
        self._save_settings()
        # Поднять туннель к VPS, если включена публикация в интернет
        if self.smpp_tunnel_enabled.isChecked():
            self._start_tunnel(port)
        self._update_status_bar()

    def _start_tunnel(self, port):
        import tunnel
        key = tunnel.find_tunnel_key()
        if not key:
            self._append_log("smpp", "Ключ туннеля не найден рядом с программой — публикация недоступна")
            self.smpp_tunnel_status.setText("Туннель: ключ не найден")
            self.smpp_tunnel_status.setStyleSheet("color: #f38ba8; font-weight: 600;")
            return
        self.tunnel = tunnel.ReverseTunnel(
            vps_host=TUNNEL_VPS_HOST, vps_user=TUNNEL_VPS_USER,
            key_path=key, remote_port=port, local_port=port,
            on_status=lambda up, msg: self.bridge.smpp_signal.emit(
                "tunnel_up" if up else "tunnel_down", msg),
        )
        self.tunnel.start()

    def _stop_smpp(self):
        if self.smpp_server:
            self.smpp_server.stop()
            self.smpp_server = None
        if self.tunnel:
            self.tunnel.stop()
            self.tunnel = None
            self.smpp_tunnel_status.setText("Туннель: выключен")
            self.smpp_tunnel_status.setStyleSheet("color: #6c7086; font-weight: 600;")
        self.btn_smpp_start.setEnabled(True)
        self.btn_smpp_stop.setEnabled(False)
        self.smpp_status_label.setText("Сервер остановлен")
        self.smpp_status_label.setStyleSheet("color: #f38ba8; font-weight: 700; font-size: 14px;")
        self._update_status_bar()

    def _smpp_callback(self, event, data):
        self.bridge.smpp_signal.emit(event, data)

    def _smpp_status(self, event, data):
        if event == "server_started":
            self.smpp_status_label.setText(f"Сервер запущен: {data}")
            self.smpp_status_label.setStyleSheet("color: #a6e3a1; font-weight: 700; font-size: 14px;")
            self.smpp_conn_info.setText(f"Клиент должен подключаться к: {get_local_ip()}:{self.smpp_port.value()}")
            self._append_log("smpp", f"Сервер запущен на {data}")
        elif event == "server_stopped":
            self._append_log("smpp", "Сервер остановлен")
        elif event == "server_error":
            self._append_log("smpp", f"Ошибка сервера: {data}")
        elif event == "session_bound":
            self._append_log("smpp", f"Клиент привязан: {data}")
        elif event == "session_closed":
            self._append_log("smpp", f"Сессия закрыта: {data}")
        elif event == "sms_received":
            self._append_log("smpp", f"SMS принята: {data.get('to', '?')}: {data.get('message', '')[:50]}")
        elif event == "sms_sent":
            ok = data.get("success")
            mark = "✓ отправлена" if ok else f"✗ ошибка: {data.get('info')}"
            self._append_log("smpp", f"SMS {data.get('to','?')} — {mark}")
        elif event == "tunnel_up":
            domain = self.smpp_domain.text().strip() or TUNNEL_DOMAIN
            addr = f"{domain}:{self.smpp_port.value()}"
            self.smpp_tunnel_status.setText(f"Туннель: активен · клиент подключается к {addr}")
            self.smpp_tunnel_status.setStyleSheet("color: #a6e3a1; font-weight: 600;")
            self.smpp_conn_info.setText(f"Клиент должен подключаться к: {addr}")
            self._append_log("smpp", data)
        elif event == "tunnel_down":
            self.smpp_tunnel_status.setText("Туннель: переподключение…")
            self.smpp_tunnel_status.setStyleSheet("color: #f9e2af; font-weight: 600;")
            self._append_log("smpp", data)
        self._update_status_bar()

    # --- CALLS ---

    def _load_call_numbers(self):
        path, _ = QFileDialog.getOpenFileName(self, "Загрузить номера", "", "Files (*.csv *.txt *.xlsx)")
        if not path:
            return
        numbers = self._parse_numbers_file(path)
        self.call_numbers = numbers
        self.call_total = len(numbers)
        self.call_numbers_label.setText(f"Загружено: {len(numbers)} номеров (дубликаты удалены)")
        self._populate_call_numbers_table(numbers)
        self.call_success_table.setRowCount(0)
        self._reset_call_stats()
        self._update_call_stats()
        self._auto_recall_attempts = 0
        self._append_log("modem", f"Загружено {len(numbers)} номеров для обзвона")

    # Статус номера в таблице «Номера» (это единственный источник правды по обзвону)
    ANSWERED_STATUS = "Отвечен"

    def _add_number_row(self, number, status="В очереди"):
        r = self.call_numbers_list.rowCount()
        self.call_numbers_list.insertRow(r)
        self._add_check_item(self.call_numbers_list, r)
        item = QTableWidgetItem(number)
        self.call_numbers_list.setItem(r, 1, item)
        self._number_items[number] = item
        self._set_status_item(r, status)

    def _populate_call_numbers_table(self, numbers):
        """
        Быстрое заполнение таблицы «Номера» для больших баз (десятки-сотни тысяч
        строк): один setRowCount вместо insertRow на каждую строку + отключение
        перерисовки на время загрузки — иначе интерфейс зависает на минуты.
        """
        table = self.call_numbers_list
        table.setUpdatesEnabled(False)
        table.setSortingEnabled(False)
        table.setRowCount(len(numbers))
        self._number_items = {}
        for r, n in enumerate(numbers):
            self._add_check_item(table, r)
            item = QTableWidgetItem(n)
            table.setItem(r, 1, item)
            self._number_items[n] = item
            self._set_status_item(r, "В очереди")
        table.setUpdatesEnabled(True)

    def _set_status_item(self, row, status):
        item = QTableWidgetItem(status)
        colors = {
            "Отвечен": "#a6e3a1", "Звоним…": "#89b4fa", "В очереди": "#6c7086",
            "Нет ответа": "#f38ba8", "Сброшен": "#f9e2af", "Занято": "#f9e2af",
            "Отклонён": "#f38ba8", "Нет соединения": "#f38ba8", "Нет линии": "#f38ba8",
        }
        item.setForeground(QColor(colors.get(status, "#f38ba8")))
        self.call_numbers_list.setItem(row, 2, item)

    def _set_number_status(self, number, status):
        item = self._number_items.get(number)
        if item is not None:
            self._set_status_item(item.row(), status)

    def _remove_number_row(self, number):
        item = self._number_items.pop(number, None)
        if item is not None:
            self.call_numbers_list.removeRow(item.row())

    def _load_call_audio(self):
        path, _ = QFileDialog.getOpenFileName(self, "Выбрать аудио", "", "Audio (*.wav *.mp3)")
        if not path:
            return
        # Конвертацию в PCM 8кГц делает бэкенд (audio.py). Здесь только проверяем,
        # что файл читается: WAV — всегда, mp3 — только если рядом есть ffmpeg.
        import audio
        ok, msg = audio.probe_audio_support(path)
        if not ok:
            QMessageBox.warning(
                self, "Аудио не подходит",
                f"{msg}\n\nСовет: используйте WAV — он работает без доп. программ."
            )
            return
        self.call_audio_path = path
        self.call_audio_label.setText(f"Аудио: {os.path.basename(path)}  ({msg})")

    def _unanswered_numbers(self):
        """Номера, которым ещё нужно позвонить (все, кроме уже ответивших)."""
        pending = []
        for r in range(self.call_numbers_list.rowCount()):
            num_it = self.call_numbers_list.item(r, 1)
            st_it = self.call_numbers_list.item(r, 2)
            if num_it and (not st_it or st_it.text() != self.ANSWERED_STATUS):
                pending.append(num_it.text())
        return pending

    def _start_calls(self):
        self._auto_recall_attempts = 0
        self._launch_calls(self._unanswered_numbers())

    def _call_unanswered(self):
        self._auto_recall_attempts = 0
        self._launch_calls(self._unanswered_numbers())

    def _export_remaining_numbers(self):
        rows = self.call_numbers_list.rowCount()
        if rows == 0:
            QMessageBox.information(self, "Пусто", "В списке «Номера» нет номеров для экспорта.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить оставшиеся номера", "оставшиеся_номера.csv",
            "CSV (*.csv);;Текст (*.txt)")
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8-sig") as f:
                f.write("Номер;Статус\n")
                for r in range(rows):
                    num = self.call_numbers_list.item(r, 1)
                    st = self.call_numbers_list.item(r, 2)
                    f.write(f"{num.text() if num else ''};{st.text() if st else ''}\n")
            QMessageBox.information(self, "Готово", f"Сохранено номеров: {rows}\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось сохранить: {e}")

    def _launch_calls(self, numbers, auto=False):
        if self.call_handler.active:
            if not auto:
                QMessageBox.information(self, "Идёт обзвон", "Дождитесь окончания текущего обзвона или нажмите «Остановить».")
            return
        if not self.call_numbers:
            if not auto:
                QMessageBox.warning(self, "Ошибка", "Загрузите базу номеров")
            return
        if not self.call_audio_path:
            if not auto:
                QMessageBox.warning(self, "Ошибка", "Выберите аудиофайл")
            return
        if not numbers:
            if not auto:
                QMessageBox.information(self, "Готово", "Все номера уже обзвонены (ответившие пропускаются).")
            return
        voice_modems = self.modem_manager.get_active_voice_modems()
        if not voice_modems:
            if not auto:
                QMessageBox.warning(self, "Ошибка", "Нет активных модемов с поддержкой голоса (E173).\n\nПодключите модем E173 на вкладке «Модемы» и отметьте его галочкой.")
            return
        self._manual_stop = False
        # помечаем обзваниваемые как «в очереди»
        for n in numbers:
            self._set_number_status(n, "В очереди")
        self.call_progress.setVisible(True)
        self.call_progress.setMaximum(len(numbers))
        self.call_progress.setValue(0)
        self.btn_start_calls.setEnabled(False)
        self.btn_stop_calls.setEnabled(True)
        self.call_handler.start_campaign(numbers, self.call_audio_path, self.call_timeout.value())

    def _stop_calls(self):
        self._manual_stop = True
        self.call_handler.stop_campaign()
        # прерванные «Звоним…» возвращаем в «В очереди»
        for r in range(self.call_numbers_list.rowCount()):
            st = self.call_numbers_list.item(r, 2)
            if st and st.text() == "Звоним…":
                self._set_status_item(r, "В очереди")
        self.btn_start_calls.setEnabled(True)
        self.btn_stop_calls.setEnabled(False)
        self._update_call_stats()

    def _trigger_auto_recall(self):
        numbers = self._unanswered_numbers()
        if numbers and not self.call_handler.active:
            self._append_log("event", f"Авто-перезвон: попытка {self._auto_recall_attempts}/3, номеров: {len(numbers)}")
            self._launch_calls(numbers, auto=True)

    def _call_callback(self, event, data):
        if event == "calling":
            if isinstance(data, dict):
                self.bridge.call_calling_signal.emit(data.get("number", ""))
                self.bridge.log_signal.emit("event", f"Звоню {data.get('number')} через {data.get('modem')}")
            else:
                self.bridge.log_signal.emit("event", str(data))
        elif event == "progress":
            self.bridge.progress_signal.emit(f"call:{data}")
        elif event == "result":
            self.bridge.call_result_signal.emit(data)
            if isinstance(data, dict):
                mark = "✓" if data.get("success") else "✗"
                self.bridge.log_signal.emit(
                    "event", f"{mark} {data.get('number')} [{data.get('modem')}]: {data.get('status')}")
        elif event == "error":
            self.bridge.log_signal.emit("event", f"Ошибка: {data}")
            self.bridge.call_error_signal.emit(str(data))
        elif event == "modem_disabled":
            # Один модем отвалился, но кампания продолжается на остальных —
            # только запись в лог, без модалки и без сброса кнопок Начать/Стоп.
            self.bridge.log_signal.emit("event", f"⚠ {data}")
        elif event == "done":
            self.bridge.log_signal.emit("event", f"— {data} —")
            self.bridge.progress_signal.emit("call:done")

    def _on_call_error(self, msg):
        QMessageBox.warning(self, "Ошибка", msg)
        self.btn_start_calls.setEnabled(True)
        self.btn_stop_calls.setEnabled(False)

    def _on_sms_error(self, msg):
        QMessageBox.warning(self, "Ошибка", msg)
        self.btn_start_sms.setEnabled(True)
        self.btn_stop_sms.setEnabled(False)

    def _on_call_result(self, data):
        number = data.get("number", "?")
        status = data.get("status", "?")
        modem = data.get("modem", "?")
        success = data.get("success", False)
        audio_played = data.get("audio_played", 0)
        audio_total = data.get("audio_total", 0.0)

        if success:
            # ответивший уходит из «Номера» и попадает в «Успешные»
            self._remove_number_row(number)
            exists = any(self.call_success_table.item(r, 1) and self.call_success_table.item(r, 1).text() == number
                         for r in range(self.call_success_table.rowCount()))
            if not exists:
                r = self.call_success_table.rowCount()
                self.call_success_table.insertRow(r)
                self._add_check_item(self.call_success_table, r)
                self.call_success_table.setItem(r, 1, QTableWidgetItem(number))
                self.call_success_table.setItem(r, 2, QTableWidgetItem(modem))
                self.call_success_table.setItem(r, 3, QTableWidgetItem(f"{audio_total:.0f} сек"))
                if audio_total > 0:
                    pct = min(100, (audio_played / audio_total) * 100)
                    pct_item = QTableWidgetItem(f"{pct:.0f}%")
                    pct_item.setForeground(QColor("#a6e3a1") if pct >= 80 else
                                           QColor("#f9e2af") if pct >= 50 else QColor("#f38ba8"))
                else:
                    pct_item = QTableWidgetItem("?")
                self.call_success_table.setItem(r, 4, pct_item)
                self.call_success_table.setItem(r, 5, QTableWidgetItem(time.strftime("%H:%M:%S")))
                self.call_success_seconds.append(audio_played)
        else:
            self._set_number_status(number, status)

        self._update_call_stats()

    # --- Чекбоксы выбора/удаления строк ---

    def _add_check_item(self, table, row, checked=False):
        # Встроенный чекбокс самого QTableWidgetItem — НЕ QWidget/QCheckBox.
        # На больших списках (десятки-сотни тысяч строк) виджет на ячейку
        # катастрофически медленный (минуты вместо секунд), это и вызывало
        # зависание интерфейса при импорте базы в 120 000 номеров.
        item = QTableWidgetItem()
        item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        item.setCheckState(Qt.Checked if checked else Qt.Unchecked)
        table.setItem(row, 0, item)
        return item

    def _row_checked(self, table, r):
        item = table.item(r, 0)
        return bool(item and item.checkState() == Qt.Checked)

    def _set_all_checks(self, table, state):
        cs = Qt.Checked if state else Qt.Unchecked
        for r in range(table.rowCount()):
            item = table.item(r, 0)
            if item:
                item.setCheckState(cs)

    def _delete_checked(self, table, is_numbers=False):
        rows = [r for r in range(table.rowCount()) if self._row_checked(table, r)]
        if is_numbers and rows:
            removed_texts = set()
            for r in rows:
                item = table.item(r, 1)
                if item:
                    removed_texts.add(item.text())
            if removed_texts:
                self.call_numbers = [n for n in self.call_numbers if n not in removed_texts]
                for t in removed_texts:
                    self._number_items.pop(t, None)
        table.setUpdatesEnabled(False)
        for r in reversed(rows):
            table.removeRow(r)
        table.setUpdatesEnabled(True)
        if is_numbers:
            self.call_numbers_label.setText(f"Осталось: {len(self.call_numbers)} номеров")
        self._update_call_stats()
        return len(rows)

    def _reset_call_stats(self):
        self.call_success_seconds = []
        self.stat_processed.setText("0")
        self.stat_answered.setText("0")
        self.stat_avg.setText("0 сек")
        self.stat_percent.setText("0%")

    def _update_call_stats(self):
        total = self.call_total or self.call_numbers_list.rowCount()
        answered = self.call_success_table.rowCount()
        # в «Номера» остались неотвеченные; считаем те, что уже обработаны (не в очереди)
        failed = 0
        for r in range(self.call_numbers_list.rowCount()):
            st = self.call_numbers_list.item(r, 2)
            if st and st.text() not in ("В очереди", "Звоним…"):
                failed += 1
        processed = answered + failed
        avg = (sum(self.call_success_seconds) / len(self.call_success_seconds)) if self.call_success_seconds else 0
        pct = (processed / total * 100) if total else 0
        self.stat_processed.setText(f"{processed}/{total}")
        self.stat_answered.setText(str(answered))
        self.stat_avg.setText(f"{avg:.0f} сек")
        self.stat_percent.setText(f"{pct:.0f}%")

    # --- SMS ---

    def _load_sms_numbers(self):
        path, _ = QFileDialog.getOpenFileName(self, "Загрузить номера", "", "Files (*.csv *.txt *.xlsx)")
        if not path:
            return
        numbers = self._parse_numbers_file(path)
        self.sms_numbers = numbers
        self.sms_numbers_label.setText(f"Загружено: {len(numbers)} номеров (дубликаты удалены)")
        self.sms_numbers_list.setRowCount(0)
        for n in numbers:
            r = self.sms_numbers_list.rowCount()
            self.sms_numbers_list.insertRow(r)
            self.sms_numbers_list.setItem(r, 0, QTableWidgetItem(n))
        self.sms_success_table.setRowCount(0)
        self.sms_failed_table.setRowCount(0)

    def _update_sms_counter(self):
        text = self.sms_message.toPlainText()
        is_unicode = any(ord(c) > 127 for c in text)
        max_chars = 70 if is_unicode else 160
        count = len(text)
        parts = (count // max_chars) + 1 if count % max_chars else (count // max_chars if count else 0)
        if parts == 0:
            parts = 1
        encoding = "UCS-2 (кириллица)" if is_unicode else "GSM (латиница)"
        self.sms_counter.setText(f"{count} символов, {parts} SMS, {encoding}")

    def _start_sms_campaign(self):
        if not self.sms_numbers:
            QMessageBox.warning(self, "Ошибка", "Загрузите базу номеров")
            return
        message = self.sms_message.toPlainText()
        if not message:
            QMessageBox.warning(self, "Ошибка", "Введите текст сообщения")
            return
        modems = self.modem_manager.get_active_modems()
        if not modems:
            QMessageBox.warning(self, "Ошибка", "Нет активных модемов.\n\nПодключите модем на вкладке «Модемы» и отметьте его галочкой.")
            return
        self.sms_progress.setVisible(True)
        self.sms_progress.setMaximum(len(self.sms_numbers))
        self.sms_progress.setValue(0)
        self.sms_success_table.setRowCount(0)
        self.sms_failed_table.setRowCount(0)
        self.sms_results = []
        self.btn_start_sms.setEnabled(False)
        self.btn_stop_sms.setEnabled(True)
        self.sms_handler.start_campaign(self.sms_numbers, message)

    def _stop_sms_campaign(self):
        self.sms_handler.stop_campaign()
        self.btn_start_sms.setEnabled(True)
        self.btn_stop_sms.setEnabled(False)

    def _sms_callback(self, event, data):
        if event == "sending":
            self.bridge.log_signal.emit("event", str(data))
        elif event == "progress":
            self.bridge.progress_signal.emit(f"sms:{data}")
        elif event == "result":
            self.bridge.sms_result_signal.emit(data)
            if isinstance(data, dict):
                mark = "✓" if data.get("success") else "✗"
                self.bridge.log_signal.emit(
                    "event", f"{mark} SMS {data.get('number')} [{data.get('modem')}]: {data.get('status')}")
        elif event == "error":
            self.bridge.log_signal.emit("event", f"Ошибка: {data}")
            self.bridge.sms_error_signal.emit(str(data))
        elif event == "done":
            self.bridge.log_signal.emit("event", f"— {data} —")
            self.bridge.progress_signal.emit("sms:done")

    def _on_sms_result(self, data):
        number = data.get("number", "?")
        status = data.get("status", "?")
        modem = data.get("modem", "?")
        success = data.get("success", False)

        if success:
            r = self.sms_success_table.rowCount()
            self.sms_success_table.insertRow(r)
            self.sms_success_table.setItem(r, 0, QTableWidgetItem(number))
            self.sms_success_table.setItem(r, 1, QTableWidgetItem(modem))
            self.sms_success_table.setItem(r, 2, QTableWidgetItem("Отправлено"))
        else:
            r = self.sms_failed_table.rowCount()
            self.sms_failed_table.insertRow(r)
            self.sms_failed_table.setItem(r, 0, QTableWidgetItem(number))
            self.sms_failed_table.setItem(r, 1, QTableWidgetItem(modem))
            self.sms_failed_table.setItem(r, 2, QTableWidgetItem(status))

    # --- PROGRESS ---

    def _update_progress(self, data):
        if data.startswith("call:"):
            msg = data[5:]
            if msg == "done":
                self.btn_start_calls.setEnabled(True)
                self.btn_stop_calls.setEnabled(False)
                self.call_status_label.setText("Обзвон завершён")
                self.call_sub_tabs.setCurrentIndex(1)
                if (self.call_auto_recall.isChecked() and not self._manual_stop
                        and self._auto_recall_attempts < 3 and self._unanswered_numbers()):
                    self._auto_recall_attempts += 1
                    self.call_status_label.setText(
                        f"Обзвон завершён. Авто-перезвон неотвеченным через минуту "
                        f"(попытка {self._auto_recall_attempts}/3)")
                    QTimer.singleShot(60000, self._trigger_auto_recall)
            else:
                self.call_status_label.setText(msg)
                val = self.call_progress.value() + 1
                self.call_progress.setValue(val)
        elif data.startswith("sms:"):
            msg = data[4:]
            if msg == "done":
                self.btn_start_sms.setEnabled(True)
                self.btn_stop_sms.setEnabled(False)
                self.sms_status_label.setText("Рассылка завершена")
                self.sms_sub_tabs.setCurrentIndex(1)
            else:
                self.sms_status_label.setText(msg)
                val = self.sms_progress.value() + 1
                self.sms_progress.setValue(val)

    # --- UTILS ---

    def _parse_numbers_file(self, path):
        numbers = []
        if path.endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(path)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell and str(cell).strip():
                            num = str(cell).strip()
                            if num.startswith('+') or num.isdigit():
                                numbers.append(num)
            except Exception as e:
                QMessageBox.warning(self, "Ошибка", f"Не удалось прочитать Excel: {e}")
        else:
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    for line in f:
                        parts = line.strip().split(',')
                        for p in parts:
                            p = p.strip()
                            if p and (p.startswith('+') or p.isdigit()):
                                numbers.append(p)
            except Exception as e:
                QMessageBox.warning(self, "Ошибка", f"Не удалось прочитать файл: {e}")
        seen = set()
        unique = []
        for n in numbers:
            if n not in seen:
                seen.add(n)
                unique.append(n)
        return unique

    def _refresh_sms_log(self):
        rows = self.database.get_sms_log(200)
        self.sms_log_table.setRowCount(0)
        for row in rows:
            r = self.sms_log_table.rowCount()
            self.sms_log_table.insertRow(r)
            for i, val in enumerate(row):
                self.sms_log_table.setItem(r, i, QTableWidgetItem(str(val)))

    def _refresh_call_log(self):
        rows = self.database.get_call_log(200)
        self.call_log_table.setRowCount(0)
        for row in rows:
            r = self.call_log_table.rowCount()
            self.call_log_table.insertRow(r)
            for i, val in enumerate(row):
                self.call_log_table.setItem(r, i, QTableWidgetItem(str(val)))

    def _load_settings(self):
        host = self.database.get_setting("smpp_host", "0.0.0.0")
        port = self.database.get_setting("smpp_port", "2775")
        sid = self.database.get_setting("smpp_system_id", "smppuser")
        pwd = self.database.get_setting("smpp_password", "smpppass")
        stype = self.database.get_setting("smpp_system_type", "SMPP")
        self.smpp_host.setText(host)
        self.smpp_port.setValue(int(port))
        self.smpp_system_id.setText(sid)
        self.smpp_password.setText(pwd)
        self.smpp_system_type.setText(stype)
        domain = self.database.get_setting("smpp_domain", TUNNEL_DOMAIN)
        self.smpp_domain.setText(domain)
        tun = self.database.get_setting("smpp_tunnel_enabled", "0")
        self.smpp_tunnel_enabled.setChecked(tun == "1")

    def _save_settings(self):
        self.database.set_setting("smpp_host", self.smpp_host.text())
        self.database.set_setting("smpp_port", self.smpp_port.value())
        self.database.set_setting("smpp_system_id", self.smpp_system_id.text())
        self.database.set_setting("smpp_password", self.smpp_password.text())
        self.database.set_setting("smpp_system_type", self.smpp_system_type.text())
        self.database.set_setting("smpp_domain", self.smpp_domain.text())
        self.database.set_setting("smpp_tunnel_enabled", "1" if self.smpp_tunnel_enabled.isChecked() else "0")

    def _update_status_bar(self):
        modems = self.modem_manager.get_all_status()
        connected = sum(1 for m in modems if m['connected'])
        active = sum(1 for m in modems if m['connected'] and m.get('active', True))
        voice = sum(1 for m in modems if m['connected'] and m['supports_voice'] and m.get('active', True))
        smpp_status = "SMPP: запущен" if self.smpp_server and self.smpp_server.is_running() else "SMPP: остановлен"
        if self.smpp_server and self.smpp_server.is_running():
            sessions = self.smpp_server.session_count()
            smpp_status = f"SMPP: запущен ({sessions} клиент(ов))"
        smpp_color = "#a6e3a1" if (self.smpp_server and self.smpp_server.is_running()) else "#6c7086"
        self.status_bar.showMessage(f"  Модемы: {connected} подключено, {active} активно ({voice} с голосом)   |   {smpp_status}  ")
        self.status_bar.setStyleSheet(f"""
            QStatusBar {{
                background-color: #181825;
                color: {smpp_color};
                border-top: 1px solid #313244;
                font-weight: 600;
                padding: 4px;
            }}
        """)

    def closeEvent(self, event):
        reply = QMessageBox.question(
            self, "Выход",
            "Закрыть программу? Активные обзвоны и SMPP-сервер будут остановлены.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply != QMessageBox.Yes:
            event.ignore()
            return
        self.modem_manager.disconnect_all()
        if self.smpp_server:
            self.smpp_server.stop()
        if self.tunnel:
            self.tunnel.stop()
        event.accept()


def _setup_crash_logging():
    """
    При console=False необработанные исключения и предупреждения раньше пропадали
    бесследно — при зависании/сбое не оставалось никакого следа для диагностики.
    Теперь всё пишется в app.log рядом с программой, а необработанное исключение
    (в т.ч. внутри цикла событий Qt) не просто убивает процесс молча — попадает в лог.
    """
    base = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.getcwd()
    log_path = os.path.join(base, "app.log")
    logging.basicConfig(
        filename=log_path, level=logging.INFO, encoding="utf-8",
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    )

    def _excepthook(exc_type, exc_value, exc_tb):
        logging.critical("Необработанное исключение", exc_info=(exc_type, exc_value, exc_tb))

    sys.excepthook = _excepthook
    return log_path


def main():
    _setup_crash_logging()
    logging.info("=== SMPP Gateway запущен ===")
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    app.setStyleSheet(STYLESHEET)
    font = QFont("Segoe UI", 10)
    app.setFont(font)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
